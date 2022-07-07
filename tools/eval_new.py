import json
import os
from pycocotools.coco import COCO
from cocoeval_adjusted import COCOeval
import openpyxl

from eval_filter_new import filter_eval


def read_json(filename="annotations//annotations.json"):
    if os.stat(filename).st_size == 0:
        return []
    else:
        with open(filename, "r") as f:
            json_file = json.load(f)
            return json_file


annotations = read_json("annotations//annotations.json")["annotations"]


def evaluate_results(kps_data, results_file):
    res = kps_data.loadRes(results_file)
    imgids = sorted(kps_data.getImgIds())
    coco_eval = COCOeval(cocoGt=kps_data, cocoDt=res, iouType="keypoints")
    coco_eval.params.imgIds = imgids
    coco_eval.evaluate()
    coco_eval.accumulate()
    results = coco_eval.summarize()
    return results


def write_excel(model_name, input_size, data_subset, data_variable, ms, results):
    try:
        wb = openpyxl.load_workbook(
            "results_new_labels.xlsx")
    except:
        wb = openpyxl.Workbook()

        ws1 = wb.create_sheet("Full", 1)
        ws1.title = "Full"
        ws2 = wb.create_sheet("Occlusion", 2)
        ws2.title = "Occlusion"
        ws3 = wb.create_sheet("Occlusion type", 3)
        ws3.title = "Occlusion type"
        ws4 = wb.create_sheet("Truncation", 4)
        ws4.title = "Truncation"
        ws5 = wb.create_sheet("Resolution", 5)
        ws5.title = "Resolution"
        ws6 = wb.create_sheet("Grouping", 6)
        ws6.title = "Grouping"
        ws7 = wb.create_sheet("Image size", 7)
        ws7.title = "Image size"
    titles = ["Method", "Input size", "Subset", "Variable", "Multi-scale", "AP", "AP50", "AP75", "APM", "APL", "AR",
              "AR50",
              "AR75", "ARM", "ARL"]
    check = True
    sheet = wb[data_subset]
    row2 = 0
    max_row = sheet.max_row
    for row in range(0, sheet.max_row + 1):
        row2 += 1
        if sheet.cell(row=row2, column=1).value == model_name and sheet.cell(row=row2, column=2).value == input_size and \
                sheet.cell(row=row2, column=3).value == data_subset:
            if data_variable == "-":
                check = False
                break
            elif str(data_variable) == str(sheet.cell(row=row2, column=4).value):
                check = False
                break
        elif row == 0 and str(sheet.cell(row=row2, column=1).value) != "Method":
            for column in range(0, len(titles)):
                c = sheet.cell(row=row + 1, column=column + 1)
                c.value = titles[column]
        elif int(row + 1) == int(max_row + 1):
            for column in range(0, len(titles)):
                c = sheet.cell(row=row2, column=column + 1)
                if column == 0:
                    c.value = model_name
                elif column == 1:
                    c.value = input_size
                elif column == 2:
                    c.value = data_subset
                elif column == 3:
                    c.value = data_variable
                elif column == 4:
                    c.value = ms
                else:
                    c.value = results[column - 5]
    if check:
        wb.save("results_new_labels.xlsx")


def run_tests_models(model_name, input_size, ms, model_results):
    filter_eval(0, 2, coco_val2017, annotations)
    res = evaluate_results(kps_data=COCO("coco_val2017.json"), results_file=model_results)
    write_excel(model_name, input_size, "Occlusion", "Visible", ms, res)

    filter_eval(0, 1, coco_val2017, annotations)
    res = evaluate_results(kps_data=COCO("coco_val2017.json"), results_file=model_results)
    write_excel(model_name, input_size, "Occlusion", "Occluded", ms, res)

    filter_eval(0, 0, coco_val2017, annotations)
    res = evaluate_results(kps_data=COCO("coco_val2017.json"), results_file=model_results)
    write_excel(model_name, input_size, "Full", "-", ms, res)

    filter_eval(1, 1, coco_val2017, annotations)
    res = evaluate_results(kps_data=COCO("coco_val2017.json"), results_file=model_results)
    write_excel(model_name, input_size, "Occlusion type", "Self", ms, res)

    filter_eval(1, 2, coco_val2017, annotations)
    res = evaluate_results(kps_data=COCO("coco_val2017.json"), results_file=model_results)
    write_excel(model_name, input_size, "Occlusion type", "Person", ms, res)

    filter_eval(1, 3, coco_val2017, annotations)
    res = evaluate_results(kps_data=COCO("coco_val2017.json"), results_file=model_results)
    write_excel(model_name, input_size, "Occlusion type", "Environment", ms, res)

    filter_eval(2, 0, coco_val2017, annotations)
    res = evaluate_results(kps_data=COCO("coco_val2017.json"), results_file=model_results)
    write_excel(model_name, input_size, "Truncation", "keyp==0", ms, res)

    filter_eval(2, 1, coco_val2017, annotations)
    res = evaluate_results(kps_data=COCO("coco_val2017.json"), results_file=model_results)
    write_excel(model_name, input_size, "Truncation", "5>keyp>0", ms, res)

    filter_eval(2, 2, coco_val2017, annotations)
    res = evaluate_results(kps_data=COCO("coco_val2017.json"), results_file=model_results)
    write_excel(model_name, input_size, "Truncation", "5<=keyp<9", ms, res)

    filter_eval(2, 3, coco_val2017, annotations)
    res = evaluate_results(kps_data=COCO("coco_val2017.json"), results_file=model_results)
    write_excel(model_name, input_size, "Truncation", "9<=keyp<13", ms, res)

    filter_eval(2, 4, coco_val2017, annotations)
    res = evaluate_results(kps_data=COCO("coco_val2017.json"), results_file=model_results)
    write_excel(model_name, input_size, "Truncation", "13<=keyp", ms, res)

    filter_eval(2, 5, coco_val2017, annotations)
    res = evaluate_results(kps_data=COCO("coco_val2017.json"), results_file=model_results)
    write_excel(model_name, input_size, "Truncation", "keyp>0", ms, res)

    filter_eval(3, 1, coco_val2017, annotations)
    res = evaluate_results(kps_data=COCO("coco_val2017.json"), results_file=model_results)
    write_excel(model_name, input_size, "Resolution", "area<32^2", ms, res)

    filter_eval(3, 2, coco_val2017, annotations)
    res = evaluate_results(kps_data=COCO("coco_val2017.json"), results_file=model_results)
    write_excel(model_name, input_size, "Resolution", "32^2<area<96^2", ms, res)

    filter_eval(3, 3, coco_val2017, annotations)
    res = evaluate_results(kps_data=COCO("coco_val2017.json"), results_file=model_results)
    write_excel(model_name, input_size, "Resolution", "96^2<area", ms, res)

    filter_eval(4, 0, coco_val2017, annotations)
    res = evaluate_results(kps_data=COCO("coco_val2017.json"), results_file=model_results)
    write_excel(model_name, input_size, "Grouping", "Head", ms, res)

    filter_eval(4, 1, coco_val2017, annotations)
    res = evaluate_results(kps_data=COCO("coco_val2017.json"), results_file=model_results)
    write_excel(model_name, input_size, "Grouping", "Core body", ms, res)

    filter_eval(4, 2, coco_val2017, annotations)
    res = evaluate_results(kps_data=COCO("coco_val2017.json"), results_file=model_results)
    write_excel(model_name, input_size, "Grouping", "Arms", ms, res)

    filter_eval(4, 3, coco_val2017, annotations)
    res = evaluate_results(kps_data=COCO("coco_val2017.json"), results_file=model_results)
    write_excel(model_name, input_size, "Grouping", "Legs", ms, res)


annotations_json = read_json()
coco_val2017 = 'C://Users//hugob//Downloads//Master_thesis//coco_test//annotations//person_keypoints_val2017.json'


def test_all_models(test_function):
    # import these models from github
    # https://github.com/leoxiaobin/deep-high-resolution-net.pytorch
    ResNet_152_384x288 = read_json("ResNet_152_384x288.json")  # models//ResNet_152_384x288.json
    HRNet_w48_384x288 = read_json("HRNet_w48_384x288.json")

    # https://github.com/HRNet/DEKR
    DEKR_w48_640x640_with_multi_scale = read_json("DEKR_w48_640x640_with_mutli_scale.json")

    # https://github.com/ilovepose/DarkPose
    DarkPose_w48_384x288 = read_json("DarkPose_w48_384x288.json")

    test_function("ResNet_152", "384x288", "-", ResNet_152_384x288)
    test_function("HRNet_W48", "384x288", "-", HRNet_w48_384x288)
    test_function("DEKR_w48_ms", "640x640", "v", DEKR_w48_640x640_with_multi_scale)
    test_function("DarkPose_w48", "384x288", "-", DarkPose_w48_384x288)


# test_all_models(run_tests_models)
